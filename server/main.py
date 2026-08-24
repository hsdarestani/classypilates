import csv
import io
import json
import os
import re
import secrets
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

import jwt
from dateutil import parser as date_parser
from fastapi import Cookie, Depends, FastAPI, File, HTTPException, Query, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from openpyxl import load_workbook
from passlib.context import CryptContext
from pydantic import BaseModel, EmailStr
from sqlalchemy import Boolean, DateTime, ForeignKey, Integer, String, Text, create_engine, func, inspect, select, text
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:////data/classy.db")
JWT_SECRET = os.getenv("JWT_SECRET", secrets.token_hex(32))
JWT_TTL_HOURS = int(os.getenv("JWT_TTL_HOURS", "24"))
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", "/data/uploads"))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
COACH_PHOTO_DIR = UPLOAD_DIR / "coach-photos"
COACH_PHOTO_DIR.mkdir(parents=True, exist_ok=True)
MAX_COACH_PHOTO_BYTES = 5 * 1024 * 1024

connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, pool_pre_ping=True, connect_args=connect_args)
SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)

class Base(DeclarativeBase):
    pass

class UserRole(Base):
    __tablename__ = "user_roles"
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id", ondelete="CASCADE"), primary_key=True)
    role_id: Mapped[int] = mapped_column(ForeignKey("roles.id", ondelete="CASCADE"), primary_key=True)

class Role(Base):
    __tablename__ = "roles"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(80), unique=True)
    description: Mapped[str] = mapped_column(String(255), default="")
    permissions_json: Mapped[str] = mapped_column(Text, default="[]")
    system: Mapped[bool] = mapped_column(Boolean, default=False)

    @property
    def permissions(self):
        try:
            return json.loads(self.permissions_json or "[]")
        except Exception:
            return []

class User(Base):
    __tablename__ = "users"
    id: Mapped[int] = mapped_column(primary_key=True)
    email: Mapped[str] = mapped_column(String(255), unique=True, index=True)
    password_hash: Mapped[str] = mapped_column(String(255))
    first_name: Mapped[str] = mapped_column(String(120), default="")
    last_name: Mapped[str] = mapped_column(String(120), default="")
    is_active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    roles: Mapped[list[Role]] = relationship(secondary="user_roles", lazy="selectin")
    coach: Mapped[Optional["Coach"]] = relationship(back_populates="user", uselist=False)
    customer_profile: Mapped[Optional["CustomerProfile"]] = relationship(back_populates="user", uselist=False)

class CustomerProfile(Base):
    __tablename__ = "customer_profiles"
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id", ondelete="CASCADE"), primary_key=True)
    phone: Mapped[str] = mapped_column(String(80), default="")
    birth_date: Mapped[str] = mapped_column(String(20), default="")
    emergency_contact: Mapped[str] = mapped_column(String(120), default="")
    marketing_opt_in: Mapped[bool] = mapped_column(Boolean, default=False)
    credits: Mapped[int] = mapped_column(Integer, default=0)
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    user: Mapped[User] = relationship(back_populates="customer_profile")

class Studio(Base):
    __tablename__ = "studios"
    id: Mapped[str] = mapped_column(String(50), primary_key=True)
    name: Mapped[str] = mapped_column(String(160))
    address: Mapped[str] = mapped_column(String(255), default="")
    capacity: Mapped[int] = mapped_column(Integer, default=10)

class Coach(Base):
    __tablename__ = "coaches"
    id: Mapped[int] = mapped_column(primary_key=True)
    user_id: Mapped[Optional[int]] = mapped_column(ForeignKey("users.id", ondelete="SET NULL"), unique=True, nullable=True)
    display_name: Mapped[str] = mapped_column(String(160))
    photo_url: Mapped[str] = mapped_column(String(500), default="")
    bio: Mapped[str] = mapped_column(Text, default="")
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    user: Mapped[Optional[User]] = relationship(back_populates="coach")

class ClassSession(Base):
    __tablename__ = "classes"
    id: Mapped[int] = mapped_column(primary_key=True)
    studio_id: Mapped[str] = mapped_column(ForeignKey("studios.id"), index=True)
    title: Mapped[str] = mapped_column(String(180))
    class_type: Mapped[str] = mapped_column(String(80), default="Reformer")
    coach_id: Mapped[Optional[int]] = mapped_column(ForeignKey("coaches.id", ondelete="SET NULL"), nullable=True)
    starts_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), index=True)
    duration: Mapped[int] = mapped_column(Integer, default=50)
    capacity: Mapped[int] = mapped_column(Integer, default=10)
    imported_bookings: Mapped[int] = mapped_column(Integer, default=0)
    source_bookings_total: Mapped[int] = mapped_column(Integer, default=0)
    status: Mapped[str] = mapped_column(String(30), default="active")
    created_by: Mapped[Optional[int]] = mapped_column(ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    coach: Mapped[Optional[Coach]] = relationship()
    studio: Mapped[Studio] = relationship()

class Booking(Base):
    __tablename__ = "bookings"
    id: Mapped[int] = mapped_column(primary_key=True)
    reference: Mapped[str] = mapped_column(String(40), unique=True, index=True)
    class_id: Mapped[int] = mapped_column(ForeignKey("classes.id"), index=True)
    customer_name: Mapped[str] = mapped_column(String(180), default="")
    email: Mapped[str] = mapped_column(String(255), index=True)
    phone: Mapped[str] = mapped_column(String(80), default="")
    spot_number: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    status: Mapped[str] = mapped_column(String(30), default="reserved")
    payment_status: Mapped[str] = mapped_column(String(30), default="pending")
    payment_method: Mapped[str] = mapped_column(String(60), default="")
    amount_cents: Mapped[int] = mapped_column(Integer, default=0)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    klass: Mapped[ClassSession] = relationship()

class CustomerBookingLink(Base):
    __tablename__ = "customer_booking_links"
    booking_id: Mapped[int] = mapped_column(ForeignKey("bookings.id", ondelete="CASCADE"), primary_key=True)
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id", ondelete="CASCADE"), index=True)
    linked_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))

class PublicClassMap(Base):
    __tablename__ = "public_class_map"
    external_id: Mapped[str] = mapped_column(String(80), primary_key=True)
    class_id: Mapped[int] = mapped_column(ForeignKey("classes.id", ondelete="CASCADE"), unique=True)

class Waitlist(Base):
    __tablename__ = "waitlist"
    id: Mapped[int] = mapped_column(primary_key=True)
    class_id: Mapped[int] = mapped_column(ForeignKey("classes.id"), index=True)
    email: Mapped[str] = mapped_column(String(255), index=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    klass: Mapped[ClassSession] = relationship()

class CustomerWaitlistLink(Base):
    __tablename__ = "customer_waitlist_links"
    waitlist_id: Mapped[int] = mapped_column(ForeignKey("waitlist.id", ondelete="CASCADE"), primary_key=True)
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id", ondelete="CASCADE"), index=True)
    linked_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))

class ScheduleUpload(Base):
    __tablename__ = "schedule_uploads"
    id: Mapped[int] = mapped_column(primary_key=True)
    filename: Mapped[str] = mapped_column(String(255))
    saved_path: Mapped[str] = mapped_column(String(500))
    uploaded_by: Mapped[int] = mapped_column(ForeignKey("users.id"))
    status: Mapped[str] = mapped_column(String(60), default="received")
    imported_rows: Mapped[int] = mapped_column(Integer, default=0)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))

Base.metadata.create_all(engine)


def migrate_schema():
    """Keep the lightweight deployment schema compatible without a migration service."""
    columns = {column["name"] for column in inspect(engine).get_columns("classes")}
    if "imported_bookings" not in columns:
        with engine.begin() as connection:
            connection.execute(text("ALTER TABLE classes ADD COLUMN imported_bookings INTEGER NOT NULL DEFAULT 0"))
    if "source_bookings_total" not in columns:
        with engine.begin() as connection:
            connection.execute(text("ALTER TABLE classes ADD COLUMN source_bookings_total INTEGER NOT NULL DEFAULT 0"))


migrate_schema()

pwd = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer(auto_error=False)

ALL_PERMISSIONS = [
    "dashboard.view", "finance.view", "bookings.view", "bookings.manage",
    "classes.view", "classes.create", "classes.edit", "classes.delete", "classes.edit_own",
    "coaches.view", "coaches.manage", "roles.manage", "users.manage", "schedules.upload",
    "marketing.view", "customers.view", "customers.manage", "pro.view"
]
DEFAULT_COACH_PERMS = ["dashboard.view", "bookings.view", "classes.view", "classes.create", "classes.edit_own", "schedules.upload"]

STUDIOS = [
    ("bhf1", "Bahnhofsviertel · 1F", "Kaiserstraße 61 · 60329 Frankfurt", 8),
    ("ladies", "Bahnhofsviertel · Ladies 2F", "Kaiserstraße 61 · 60329 Frankfurt", 10),
    ("sachsen", "Sachsenhausen", "Zum Gipfelhof 5 · 60594 Frankfurt", 12),
    ("bornheim", "Bornheim", "Wiesenstraße 33 · 60385 Frankfurt", 8),
    ("mid", "Mid", "Große Eschenheimer Straße 45 · 60313 Frankfurt", 10),
    ("oval", "Oval", "Baseler Straße 10 · 60329 Frankfurt", 10),
]

def db_session():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def seed():
    with SessionLocal() as db:
        system_roles = (
            ("Administrator", "Full access", ["*"]),
            ("Coach", "Standard access for coaches", DEFAULT_COACH_PERMS),
            ("Customer", "Customer account for bookings and Class Credits", []),
        )
        for name, description, permissions in system_roles:
            role = db.scalar(select(Role).where(Role.name == name))
            if role:
                role.description = description
            else:
                db.add(Role(name=name, description=description, permissions_json=json.dumps(permissions), system=True))
        for sid, name, address, cap in STUDIOS:
            studio = db.get(Studio, sid)
            if studio:
                studio.name = name
                studio.address = address
            else:
                db.add(Studio(id=sid, name=name, address=address, capacity=cap))
        db.commit()
seed()

def user_permissions(user: User):
    perms = set()
    for role in user.roles:
        perms.update(role.permissions)
    return perms

def can(user: User, permission: str):
    perms = user_permissions(user)
    return "*" in perms or permission in perms

def portal_for(user: User):
    perms = user_permissions(user)
    if "*" in perms or any(p in perms for p in ALL_PERMISSIONS):
        if user.coach and "*" not in perms and "roles.manage" not in perms and "users.manage" not in perms:
            return "/coach"
        return "/admin"
    return "/account"

def make_token(user: User):
    now = datetime.now(timezone.utc)
    return jwt.encode({"sub": str(user.id), "iat": int(now.timestamp()), "exp": int((now + timedelta(hours=JWT_TTL_HOURS)).timestamp())}, JWT_SECRET, algorithm="HS256")

def current_user(credentials: HTTPAuthorizationCredentials = Depends(security), cp_session: Optional[str] = Cookie(default=None), db: Session = Depends(db_session)):
    token = credentials.credentials if credentials else cp_session
    if not token:
        raise HTTPException(401, "login_required")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        user = db.get(User, int(payload["sub"]))
    except Exception:
        raise HTTPException(401, "invalid_token")
    if not user or not user.is_active:
        raise HTTPException(401, "inactive_user")
    return user

def optional_user(credentials: HTTPAuthorizationCredentials = Depends(security), cp_session: Optional[str] = Cookie(default=None), db: Session = Depends(db_session)):
    token = credentials.credentials if credentials else cp_session
    if not token:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        user = db.get(User, int(payload["sub"]))
    except Exception:
        return None
    return user if user and user.is_active else None

def require(permission: str):
    def dep(user: User = Depends(current_user)):
        if not can(user, permission):
            raise HTTPException(403, "permission_denied")
        return user
    return dep

def user_dict(user: User):
    return {
        "id": user.id, "email": user.email, "first_name": user.first_name, "last_name": user.last_name,
        "is_active": user.is_active,
        "roles": [{"id": r.id, "name": r.name, "permissions": r.permissions} for r in user.roles],
        "permissions": sorted(user_permissions(user)),
        "coach": None if not user.coach else {"id": user.coach.id, "display_name": user.coach.display_name, "photo_url": user.coach.photo_url},
        "portal": portal_for(user)
    }

def coach_dict(coach: Coach):
    return {
        "id": coach.id,
        "display_name": coach.display_name,
        "photo_url": coach.photo_url,
        "bio": coach.bio,
        "active": coach.active,
        "email": coach.user.email if coach.user else "",
    }

def coach_photo_format(content: bytes):
    if content.startswith(b"\xff\xd8\xff"):
        return ".jpg", "image/jpeg"
    if content.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png", "image/png"
    if len(content) >= 12 and content[:4] == b"RIFF" and content[8:12] == b"WEBP":
        return ".webp", "image/webp"
    raise HTTPException(400, "unsupported_image")

def managed_coach_photo_path(photo_url: str):
    prefix = "/api/media/coach-photos/"
    if not photo_url.startswith(prefix):
        return None
    filename = photo_url.removeprefix(prefix)
    if not re.fullmatch(r"coach-\d+-[a-f0-9]{24}\.(?:jpg|png|webp)", filename):
        return None
    path = (COACH_PHOTO_DIR / filename).resolve()
    return path if path.parent == COACH_PHOTO_DIR.resolve() else None

def remove_managed_coach_photo(photo_url: str):
    path = managed_coach_photo_path(photo_url or "")
    if path:
        path.unlink(missing_ok=True)

async def store_coach_photo(file: UploadFile, coach: Coach, db: Session):
    content = await file.read(MAX_COACH_PHOTO_BYTES + 1)
    if not content:
        raise HTTPException(400, "empty_image")
    if len(content) > MAX_COACH_PHOTO_BYTES:
        raise HTTPException(413, "image_too_large")
    extension, _ = coach_photo_format(content)
    filename = f"coach-{coach.id}-{secrets.token_hex(12)}{extension}"
    path = COACH_PHOTO_DIR / filename
    path.write_bytes(content)
    old_photo = coach.photo_url
    coach.photo_url = f"/api/media/coach-photos/{filename}"
    try:
        db.commit()
    except Exception:
        path.unlink(missing_ok=True)
        db.rollback()
        raise
    remove_managed_coach_photo(old_photo)
    return coach_dict(coach)

def class_dict(c: ClassSession, db: Session):
    live_reserved = db.scalar(select(func.count(Booking.id)).where(Booking.class_id == c.id, Booking.status == "reserved")) or 0
    imported_reserved = max(0, int(c.imported_bookings or 0))
    reserved = min(c.capacity, imported_reserved + live_reserved)
    starts_at = c.starts_at
    if starts_at.tzinfo is None:
        starts_at = starts_at.replace(tzinfo=ZoneInfo("Europe/Berlin"))
    return {
        "id": c.id, "studio": c.studio_id, "studio_name": c.studio.name if c.studio else c.studio_id,
        "name": c.title, "type": c.class_type, "coach": c.coach.display_name if c.coach else "Classy Coach",
        "coach_id": c.coach_id, "starts_at": starts_at.isoformat(), "duration": c.duration,
        "capacity": c.capacity, "reserved": reserved, "imported_reserved": imported_reserved,
        "source_bookings_total": int(c.source_bookings_total or 0),
        "live_reserved": live_reserved, "spots": max(0, c.capacity - reserved), "status": c.status
    }

app = FastAPI(title="Classy Pilates Production API", version="1.2")
app.add_middleware(CORSMiddleware, allow_origins=["https://classy.smarbiz.sbs", "http://localhost", "http://127.0.0.1"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

@app.get("/api/media/coach-photos/{filename}")
def coach_photo_media(filename: str):
    if not re.fullmatch(r"coach-\d+-[a-f0-9]{24}\.(?:jpg|png|webp)", filename):
        raise HTTPException(404, "image_not_found")
    path = (COACH_PHOTO_DIR / filename).resolve()
    if path.parent != COACH_PHOTO_DIR.resolve() or not path.is_file():
        raise HTTPException(404, "image_not_found")
    media_type = {".jpg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}[path.suffix]
    return FileResponse(path, media_type=media_type, headers={"Cache-Control": "public, max-age=31536000, immutable"})

class LoginIn(BaseModel):
    email: EmailStr
    password: str

class RegisterIn(BaseModel):
    email: EmailStr
    password: str
    first_name: str = ""
    last_name: str = ""
    phone: str = ""
    birth_date: str = ""
    emergency_contact: str = ""
    marketing_opt_in: bool = False

class BootstrapIn(BaseModel):
    email: EmailStr
    password: str
    first_name: str = "Admin"
    last_name: str = ""

class RoleIn(BaseModel):
    name: str
    description: str = ""
    permissions: list[str] = []

class StaffIn(BaseModel):
    email: EmailStr
    password: str
    first_name: str = ""
    last_name: str = ""
    role_ids: list[int] = []
    coach_name: str = ""
    coach_photo: str = ""

class StaffUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    is_active: Optional[bool] = None
    role_ids: Optional[list[int]] = None

class CoachProfileIn(BaseModel):
    display_name: str
    bio: str = ""
    active: Optional[bool] = None

class CustomerProfileIn(BaseModel):
    first_name: str = ""
    last_name: str = ""
    phone: str = ""
    birth_date: str = ""
    emergency_contact: str = ""
    marketing_opt_in: bool = False

class PasswordChangeIn(BaseModel):
    current_password: str
    new_password: str

class CoachAccountIn(BaseModel):
    email: EmailStr
    password: str
    first_name: str = ""
    last_name: str = ""

class PasswordResetIn(BaseModel):
    new_password: str

class CustomerAdminUpdate(BaseModel):
    credits: Optional[int] = None
    is_active: Optional[bool] = None

class ClassIn(BaseModel):
    studio_id: str
    title: str
    class_type: str = "Reformer"
    coach_id: Optional[int] = None
    starts_at: datetime
    duration: int = 50
    capacity: int = 10

class BookingUpdate(BaseModel):
    status: Optional[str] = None
    payment_status: Optional[str] = None
    amount_cents: Optional[int] = None

class PublicBookingIn(BaseModel):
    classId: int | str
    email: EmailStr
    firstName: str = ""
    lastName: str = ""
    phone: str = ""
    spot: Optional[int] = None
    paymentMethod: str = ""
    studioId: str = ""
    title: str = ""
    classType: str = ""
    startsAt: Optional[datetime] = None
    duration: int = 50
    capacity: int = 10
    coachName: str = ""

class WaitlistIn(BaseModel):
    classId: int
    email: EmailStr

class BookingClaimIn(BaseModel):
    reference: str

@app.get("/api/health")
def health():
    return {"ok": True, "service": "classy-production"}

@app.get("/api/auth/status")
def auth_status(db: Session = Depends(db_session)):
    return {"configured": (db.scalar(select(func.count(User.id))) or 0) > 0}

@app.post("/api/auth/bootstrap")
def bootstrap(data: BootstrapIn, response: Response, db: Session = Depends(db_session)):
    if (db.scalar(select(func.count(User.id))) or 0) > 0:
        raise HTTPException(409, "already_configured")
    if len(data.password) < 10:
        raise HTTPException(400, "password_too_short")
    role = db.scalar(select(Role).where(Role.name == "Administrator"))
    user = User(email=data.email.lower(), password_hash=pwd.hash(data.password), first_name=data.first_name, last_name=data.last_name, roles=[role])
    db.add(user); db.commit(); db.refresh(user)
    token = make_token(user)
    response.set_cookie("cp_session", token, max_age=JWT_TTL_HOURS * 3600, httponly=True, secure=True, samesite="lax", path="/")
    return {"token": token, "user": user_dict(user)}

@app.post("/api/auth/login")
def login(data: LoginIn, response: Response, db: Session = Depends(db_session)):
    email = data.email.strip().lower()
    user = db.scalar(select(User).where(func.lower(User.email) == email))
    if not user or not pwd.verify(data.password, user.password_hash):
        raise HTTPException(401, "invalid_credentials")
    if not user.is_active:
        raise HTTPException(403, "inactive_user")
    token = make_token(user)
    response.set_cookie("cp_session", token, max_age=JWT_TTL_HOURS * 3600, httponly=True, secure=True, samesite="lax", path="/")
    return {"token": token, "user": user_dict(user)}

@app.post("/api/auth/register")
def register(data: RegisterIn, response: Response, db: Session = Depends(db_session)):
    email = data.email.strip().lower()
    if len(data.password) < 8:
        raise HTTPException(400, "password_too_short")
    if len(data.first_name.strip()) < 2 or len(data.last_name.strip()) < 2:
        raise HTTPException(400, "name_required")
    if db.scalar(select(User).where(func.lower(User.email) == email)):
        raise HTTPException(409, "email_exists")
    role = db.scalar(select(Role).where(Role.name == "Customer"))
    if not role:
        raise HTTPException(503, "customer_role_missing")
    user = User(
        email=email,
        password_hash=pwd.hash(data.password),
        first_name=data.first_name.strip(),
        last_name=data.last_name.strip(),
        roles=[role],
    )
    db.add(user)
    try:
        db.flush()
    except Exception:
        db.rollback()
        raise HTTPException(409, "email_exists")
    db.add(CustomerProfile(
        user_id=user.id,
        phone=data.phone.strip(),
        birth_date=data.birth_date.strip(),
        emergency_contact=data.emergency_contact.strip(),
        marketing_opt_in=data.marketing_opt_in,
    ))
    db.commit(); db.refresh(user)
    token = make_token(user)
    response.set_cookie("cp_session", token, max_age=JWT_TTL_HOURS * 3600, httponly=True, secure=True, samesite="lax", path="/")
    return {"token": token, "user": user_dict(user)}

@app.post("/api/auth/logout")
def auth_logout(response: Response):
    response.delete_cookie("cp_session", path="/")
    return {"ok": True}

@app.get("/api/auth/me")
def me(user: User = Depends(current_user)):
    return user_dict(user)

def customer_only(user: User = Depends(current_user)):
    if portal_for(user) != "/account":
        raise HTTPException(403, "customer_access_required")
    return user

def get_customer_profile(user: User, db: Session):
    profile = db.get(CustomerProfile, user.id)
    if not profile:
        profile = CustomerProfile(user_id=user.id)
        db.add(profile); db.commit(); db.refresh(profile)
    return profile

def as_utc(value: datetime):
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)

def customer_booking_dict(booking: Booking):
    starts_at = as_utc(booking.klass.starts_at)
    cancellation_deadline = starts_at - timedelta(hours=12)
    return {
        "reference": booking.reference,
        "name": booking.klass.title,
        "type": booking.klass.class_type,
        "studio": booking.klass.studio.name,
        "studio_address": booking.klass.studio.address,
        "coach": booking.klass.coach.display_name if booking.klass.coach else "Classy Coach",
        "starts_at": starts_at.isoformat(),
        "duration": booking.klass.duration,
        "spot_number": booking.spot_number,
        "status": booking.status,
        "payment_status": booking.payment_status,
        "payment_method": booking.payment_method,
        "amount_cents": booking.amount_cents,
        "cancellation_deadline": cancellation_deadline.isoformat(),
        "can_cancel": booking.status == "reserved" and datetime.now(timezone.utc) < cancellation_deadline,
    }

@app.get("/api/customer/dashboard")
def customer_dashboard(user: User = Depends(customer_only), db: Session = Depends(db_session)):
    profile = get_customer_profile(user, db)
    rows = db.scalars(
        select(Booking)
        .join(CustomerBookingLink, CustomerBookingLink.booking_id == Booking.id)
        .join(Booking.klass)
        .where(CustomerBookingLink.user_id == user.id)
        .order_by(ClassSession.starts_at.desc())
    ).all()
    now = datetime.now(timezone.utc)
    booking_rows = [customer_booking_dict(b) for b in rows]
    upcoming = [b for b in booking_rows if as_utc(date_parser.parse(b["starts_at"])) >= now and b["status"] == "reserved"]
    waitlist_count = db.scalar(select(func.count(CustomerWaitlistLink.waitlist_id)).where(CustomerWaitlistLink.user_id == user.id)) or 0
    return {
        "profile": {"first_name": user.first_name, "last_name": user.last_name, "email": user.email},
        "credits": profile.credits,
        "upcoming_count": len(upcoming),
        "total_bookings": len(booking_rows),
        "waitlist_count": waitlist_count,
        "next_booking": upcoming[-1] if upcoming else None,
        "recent_bookings": booking_rows[:5],
    }

@app.get("/api/customer/bookings")
def customer_bookings(user: User = Depends(customer_only), db: Session = Depends(db_session)):
    rows = db.scalars(
        select(Booking)
        .join(CustomerBookingLink, CustomerBookingLink.booking_id == Booking.id)
        .where(CustomerBookingLink.user_id == user.id)
        .order_by(Booking.created_at.desc())
    ).all()
    return {"bookings": [customer_booking_dict(b) for b in rows]}

@app.post("/api/customer/bookings/claim")
def customer_claim_booking(data: BookingClaimIn, user: User = Depends(customer_only), db: Session = Depends(db_session)):
    reference = data.reference.strip().upper()
    booking = db.scalar(select(Booking).where(Booking.reference == reference, func.lower(Booking.email) == user.email.lower()))
    if not booking:
        raise HTTPException(404, "booking_not_found")
    link = db.get(CustomerBookingLink, booking.id)
    if link and link.user_id != user.id:
        raise HTTPException(409, "booking_already_linked")
    if not link:
        db.add(CustomerBookingLink(booking_id=booking.id, user_id=user.id)); db.commit()
    return {"ok": True}

@app.delete("/api/customer/bookings/{reference}")
def customer_cancel_booking(reference: str, user: User = Depends(customer_only), db: Session = Depends(db_session)):
    booking = db.scalar(
        select(Booking)
        .join(CustomerBookingLink, CustomerBookingLink.booking_id == Booking.id)
        .where(Booking.reference == reference.strip().upper(), CustomerBookingLink.user_id == user.id)
    )
    if not booking:
        raise HTTPException(404, "booking_not_found")
    if booking.status != "reserved":
        raise HTTPException(409, "booking_not_active")
    if datetime.now(timezone.utc) >= as_utc(booking.klass.starts_at) - timedelta(hours=12):
        raise HTTPException(409, "cancellation_window_closed")
    booking.status = "cancelled"; db.commit()
    return {"ok": True}

@app.get("/api/customer/waitlist")
def customer_waitlist(user: User = Depends(customer_only), db: Session = Depends(db_session)):
    rows = db.scalars(
        select(Waitlist)
        .join(CustomerWaitlistLink, CustomerWaitlistLink.waitlist_id == Waitlist.id)
        .where(CustomerWaitlistLink.user_id == user.id)
        .order_by(Waitlist.created_at.desc())
    ).all()
    return {"waitlist": [{
        "id": row.id,
        "name": row.klass.title,
        "studio": row.klass.studio.name,
        "starts_at": as_utc(row.klass.starts_at).isoformat(),
        "coach": row.klass.coach.display_name if row.klass.coach else "Classy Coach",
        "created_at": row.created_at.isoformat(),
    } for row in rows]}

@app.delete("/api/customer/waitlist/{waitlist_id}")
def customer_leave_waitlist(waitlist_id: int, user: User = Depends(customer_only), db: Session = Depends(db_session)):
    row = db.scalar(
        select(Waitlist)
        .join(CustomerWaitlistLink, CustomerWaitlistLink.waitlist_id == Waitlist.id)
        .where(Waitlist.id == waitlist_id, CustomerWaitlistLink.user_id == user.id)
    )
    if not row:
        raise HTTPException(404, "waitlist_not_found")
    db.delete(row); db.commit()
    return {"ok": True}

@app.get("/api/customer/profile")
def customer_profile(user: User = Depends(customer_only), db: Session = Depends(db_session)):
    profile = get_customer_profile(user, db)
    return {
        "first_name": user.first_name,
        "last_name": user.last_name,
        "email": user.email,
        "phone": profile.phone,
        "birth_date": profile.birth_date,
        "emergency_contact": profile.emergency_contact,
        "marketing_opt_in": profile.marketing_opt_in,
        "credits": profile.credits,
    }

@app.patch("/api/customer/profile")
def customer_profile_update(data: CustomerProfileIn, user: User = Depends(customer_only), db: Session = Depends(db_session)):
    if len(data.first_name.strip()) < 2 or len(data.last_name.strip()) < 2:
        raise HTTPException(400, "name_required")
    profile = get_customer_profile(user, db)
    user.first_name = data.first_name.strip(); user.last_name = data.last_name.strip()
    profile.phone = data.phone.strip(); profile.birth_date = data.birth_date.strip()
    profile.emergency_contact = data.emergency_contact.strip(); profile.marketing_opt_in = data.marketing_opt_in
    db.commit()
    return {"ok": True}

@app.post("/api/customer/change-password")
def customer_change_password(data: PasswordChangeIn, user: User = Depends(customer_only), db: Session = Depends(db_session)):
    if not pwd.verify(data.current_password, user.password_hash):
        raise HTTPException(400, "current_password_invalid")
    if len(data.new_password) < 8:
        raise HTTPException(400, "password_too_short")
    user.password_hash = pwd.hash(data.new_password); db.commit()
    return {"ok": True}

@app.get("/api/staff/dashboard")
def dashboard(user: User = Depends(require("dashboard.view")), db: Session = Depends(db_session)):
    now = datetime.now(timezone.utc)
    class_count = db.scalar(select(func.count(ClassSession.id)).where(ClassSession.starts_at >= now, ClassSession.status == "active")) or 0
    live_booking_count = db.scalar(select(func.count(Booking.id)).where(Booking.status == "reserved")) or 0
    imported_booking_count = db.scalar(
        select(func.coalesce(func.sum(ClassSession.imported_bookings), 0)).where(
            ClassSession.starts_at >= now, ClassSession.status == "active"
        )
    ) or 0
    booking_count = int(live_booking_count) + int(imported_booking_count)
    coach_count = db.scalar(select(func.count(Coach.id)).where(Coach.active == True)) or 0
    paid_cents = db.scalar(select(func.coalesce(func.sum(Booking.amount_cents), 0)).where(Booking.payment_status == "paid")) or 0
    today_bookings = db.scalar(select(func.count(Booking.id)).where(Booking.created_at >= now.replace(hour=0, minute=0, second=0, microsecond=0))) or 0
    return {
        "upcoming_classes": class_count, "active_bookings": booking_count,
        "imported_bookings": int(imported_booking_count), "live_bookings": int(live_booking_count),
        "coaches": coach_count, "revenue_cents": int(paid_cents), "today_bookings": today_bookings,
    }

@app.get("/api/staff/finance")
def finance(user: User = Depends(require("finance.view")), db: Session = Depends(db_session)):
    total = int(db.scalar(select(func.coalesce(func.sum(Booking.amount_cents), 0)).where(Booking.payment_status == "paid")) or 0)
    pending = int(db.scalar(select(func.coalesce(func.sum(Booking.amount_cents), 0)).where(Booking.payment_status == "pending")) or 0)
    paid_bookings = db.scalar(select(func.count(Booking.id)).where(Booking.payment_status == "paid")) or 0
    recent = db.scalars(select(Booking).order_by(Booking.created_at.desc()).limit(30)).all()
    return {"revenue_cents": total, "pending_cents": pending, "paid_bookings": paid_bookings, "rows": [{"reference": b.reference, "email": b.email, "amount_cents": b.amount_cents, "payment_status": b.payment_status, "payment_method": b.payment_method, "created_at": b.created_at.isoformat()} for b in recent]}

@app.get("/api/staff/roles")
def list_roles(user: User = Depends(current_user), db: Session = Depends(db_session)):
    if not (can(user, "roles.manage") or can(user, "users.manage")):
        raise HTTPException(403, "permission_denied")
    roles = db.scalars(select(Role).order_by(Role.name)).all()
    return {"permissions": ALL_PERMISSIONS, "roles": [{"id": r.id, "name": r.name, "description": r.description, "permissions": r.permissions, "system": r.system} for r in roles]}

@app.post("/api/staff/roles")
def create_role(data: RoleIn, user: User = Depends(require("roles.manage")), db: Session = Depends(db_session)):
    perms = [p for p in data.permissions if p in ALL_PERMISSIONS]
    role = Role(name=data.name.strip(), description=data.description, permissions_json=json.dumps(perms), system=False)
    db.add(role)
    try: db.commit()
    except Exception: db.rollback(); raise HTTPException(409, "role_exists")
    db.refresh(role)
    return {"id": role.id}

@app.patch("/api/staff/roles/{role_id}")
def update_role(role_id: int, data: RoleIn, user: User = Depends(require("roles.manage")), db: Session = Depends(db_session)):
    role = db.get(Role, role_id)
    if not role: raise HTTPException(404, "not_found")
    requested_name = data.name.strip()
    if role.system and requested_name != role.name:
        raise HTTPException(400, "system_role_name_locked")
    role.name = requested_name; role.description = data.description
    if not role.system or role.name == "Coach":
        role.permissions_json = json.dumps([p for p in data.permissions if p in ALL_PERMISSIONS])
    elif role.name == "Administrator":
        role.permissions_json = json.dumps(["*"])
    elif role.name == "Customer":
        role.permissions_json = "[]"
    db.commit(); return {"ok": True}

@app.delete("/api/staff/roles/{role_id}")
def delete_role(role_id: int, user: User = Depends(require("roles.manage")), db: Session = Depends(db_session)):
    role = db.get(Role, role_id)
    if not role: raise HTTPException(404, "not_found")
    if role.system: raise HTTPException(400, "system_role")
    db.delete(role); db.commit(); return {"ok": True}

@app.get("/api/staff/users")
def list_users(user: User = Depends(require("users.manage")), db: Session = Depends(db_session)):
    users = db.scalars(select(User).order_by(User.created_at.desc())).all()
    return {"users": [user_dict(u) for u in users if not any(role.name == "Customer" for role in u.roles)]}

@app.get("/api/staff/customers")
def list_customers(user: User = Depends(require("customers.view")), db: Session = Depends(db_session)):
    customers = [u for u in db.scalars(select(User).order_by(User.created_at.desc())).all() if any(role.name == "Customer" for role in u.roles)]
    rows = []
    for customer in customers:
        profile = get_customer_profile(customer, db)
        booking_count = db.scalar(select(func.count(Booking.id)).where(func.lower(Booking.email) == customer.email.lower())) or 0
        active_count = db.scalar(select(func.count(Booking.id)).where(func.lower(Booking.email) == customer.email.lower(), Booking.status == "reserved")) or 0
        rows.append({
            "id": customer.id,
            "first_name": customer.first_name,
            "last_name": customer.last_name,
            "email": customer.email,
            "phone": profile.phone,
            "credits": profile.credits,
            "booking_count": booking_count,
            "active_bookings": active_count,
            "is_active": customer.is_active,
            "created_at": customer.created_at.isoformat(),
        })
    return {"customers": rows}

@app.patch("/api/staff/customers/{customer_id}")
def update_customer(customer_id: int, data: CustomerAdminUpdate, user: User = Depends(require("customers.manage")), db: Session = Depends(db_session)):
    customer = db.get(User, customer_id)
    if not customer or not any(role.name == "Customer" for role in customer.roles):
        raise HTTPException(404, "customer_not_found")
    profile = get_customer_profile(customer, db)
    if data.credits is not None:
        profile.credits = max(0, data.credits)
    if data.is_active is not None:
        customer.is_active = data.is_active
    db.commit()
    return {"ok": True}

@app.post("/api/staff/users")
def create_staff(data: StaffIn, user: User = Depends(require("users.manage")), db: Session = Depends(db_session)):
    if len(data.password) < 8: raise HTTPException(400, "password_too_short")
    roles = db.scalars(select(Role).where(Role.id.in_(data.role_ids))).all() if data.role_ids else []
    u = User(email=data.email.lower(), password_hash=pwd.hash(data.password), first_name=data.first_name, last_name=data.last_name, roles=list(roles))
    db.add(u)
    try: db.flush()
    except Exception: db.rollback(); raise HTTPException(409, "email_exists")
    if data.coach_name:
        db.add(Coach(user_id=u.id, display_name=data.coach_name, photo_url=data.coach_photo))
    db.commit(); db.refresh(u)
    return user_dict(u)

@app.patch("/api/staff/users/{user_id}")
def update_staff(user_id: int, data: StaffUpdate, user: User = Depends(require("users.manage")), db: Session = Depends(db_session)):
    target = db.get(User, user_id)
    if not target: raise HTTPException(404, "not_found")
    if data.first_name is not None: target.first_name = data.first_name
    if data.last_name is not None: target.last_name = data.last_name
    if data.is_active is not None: target.is_active = data.is_active
    if data.role_ids is not None: target.roles = list(db.scalars(select(Role).where(Role.id.in_(data.role_ids))).all())
    db.commit(); return user_dict(target)

@app.post("/api/staff/change-password")
def staff_change_password(data: PasswordChangeIn, user: User = Depends(current_user), db: Session = Depends(db_session)):
    if any(role.name == "Customer" for role in user.roles) and not user.coach:
        raise HTTPException(403, "staff_account_required")
    if not pwd.verify(data.current_password, user.password_hash):
        raise HTTPException(400, "current_password_invalid")
    if len(data.new_password) < 8:
        raise HTTPException(400, "password_too_short")
    user.password_hash = pwd.hash(data.new_password)
    db.commit()
    return {"ok": True}

@app.get("/api/staff/coaches")
def list_coaches(user: User = Depends(require("coaches.view")), db: Session = Depends(db_session)):
    coaches = db.scalars(select(Coach).order_by(Coach.display_name)).all()
    return {"coaches": [coach_dict(c) for c in coaches]}

@app.post("/api/staff/coaches/{coach_id}/account")
def create_coach_account(coach_id: int, data: CoachAccountIn, user: User = Depends(require("users.manage")), db: Session = Depends(db_session)):
    coach = db.get(Coach, coach_id)
    if not coach:
        raise HTTPException(404, "coach_not_found")
    if coach.user_id:
        raise HTTPException(409, "coach_account_exists")
    if len(data.password) < 8:
        raise HTTPException(400, "password_too_short")
    coach_role = db.scalar(select(Role).where(Role.name == "Coach"))
    if not coach_role:
        raise HTTPException(500, "coach_role_missing")
    account = User(
        email=data.email.lower(), password_hash=pwd.hash(data.password),
        first_name=data.first_name.strip(), last_name=data.last_name.strip(),
        roles=[coach_role],
    )
    db.add(account)
    try:
        db.flush()
    except Exception:
        db.rollback()
        raise HTTPException(409, "email_exists")
    coach.user_id = account.id
    db.commit()
    return coach_dict(coach)

@app.post("/api/staff/coaches/{coach_id}/reset-password")
def reset_coach_password(coach_id: int, data: PasswordResetIn, user: User = Depends(require("users.manage")), db: Session = Depends(db_session)):
    coach = db.get(Coach, coach_id)
    if not coach:
        raise HTTPException(404, "coach_not_found")
    if not coach.user:
        raise HTTPException(404, "coach_account_not_found")
    if len(data.new_password) < 8:
        raise HTTPException(400, "password_too_short")
    coach.user.password_hash = pwd.hash(data.new_password)
    db.commit()
    return {"ok": True}

@app.get("/api/staff/profile")
def coach_profile(user: User = Depends(current_user)):
    if not user.coach:
        raise HTTPException(404, "coach_profile_not_found")
    return coach_dict(user.coach)

@app.patch("/api/staff/profile")
def update_coach_profile(data: CoachProfileIn, user: User = Depends(current_user), db: Session = Depends(db_session)):
    if not user.coach:
        raise HTTPException(404, "coach_profile_not_found")
    display_name = data.display_name.strip()
    if len(display_name) < 2:
        raise HTTPException(400, "coach_name_required")
    user.coach.display_name = display_name
    user.coach.bio = data.bio.strip()[:2000]
    db.commit()
    return coach_dict(user.coach)

@app.post("/api/staff/profile/photo")
async def upload_own_coach_photo(file: UploadFile = File(...), user: User = Depends(current_user), db: Session = Depends(db_session)):
    if not user.coach:
        raise HTTPException(404, "coach_profile_not_found")
    return await store_coach_photo(file, user.coach, db)

@app.delete("/api/staff/profile/photo")
def delete_own_coach_photo(user: User = Depends(current_user), db: Session = Depends(db_session)):
    if not user.coach:
        raise HTTPException(404, "coach_profile_not_found")
    old_photo = user.coach.photo_url
    user.coach.photo_url = ""
    db.commit()
    remove_managed_coach_photo(old_photo)
    return coach_dict(user.coach)

@app.patch("/api/staff/coaches/{coach_id}")
def update_coach(coach_id: int, data: CoachProfileIn, user: User = Depends(require("coaches.manage")), db: Session = Depends(db_session)):
    coach = db.get(Coach, coach_id)
    if not coach:
        raise HTTPException(404, "coach_not_found")
    display_name = data.display_name.strip()
    if len(display_name) < 2:
        raise HTTPException(400, "coach_name_required")
    coach.display_name = display_name
    coach.bio = data.bio.strip()[:2000]
    if data.active is not None:
        coach.active = data.active
    db.commit()
    return coach_dict(coach)

@app.post("/api/staff/coaches/{coach_id}/photo")
async def upload_coach_photo(coach_id: int, file: UploadFile = File(...), user: User = Depends(require("coaches.manage")), db: Session = Depends(db_session)):
    coach = db.get(Coach, coach_id)
    if not coach:
        raise HTTPException(404, "coach_not_found")
    return await store_coach_photo(file, coach, db)

@app.delete("/api/staff/coaches/{coach_id}/photo")
def delete_coach_photo(coach_id: int, user: User = Depends(require("coaches.manage")), db: Session = Depends(db_session)):
    coach = db.get(Coach, coach_id)
    if not coach:
        raise HTTPException(404, "coach_not_found")
    old_photo = coach.photo_url
    coach.photo_url = ""
    db.commit()
    remove_managed_coach_photo(old_photo)
    return coach_dict(coach)

@app.get("/api/staff/classes")
def staff_classes(user: User = Depends(require("classes.view")), db: Session = Depends(db_session)):
    q = select(ClassSession).order_by(ClassSession.starts_at.desc()).limit(300)
    rows = db.scalars(q).all()
    if user.coach and not can(user, "classes.edit"):
        rows = [r for r in rows if r.coach_id == user.coach.id]
    return {"classes": [class_dict(c, db) for c in rows]}

@app.post("/api/staff/classes")
def create_class(data: ClassIn, user: User = Depends(require("classes.create")), db: Session = Depends(db_session)):
    coach_id = data.coach_id
    if user.coach and not can(user, "classes.edit"):
        coach_id = user.coach.id
    studio = db.get(Studio, data.studio_id)
    if not studio: raise HTTPException(400, "invalid_studio")
    c = ClassSession(studio_id=data.studio_id, title=data.title, class_type=data.class_type, coach_id=coach_id, starts_at=data.starts_at, duration=max(15, data.duration), capacity=max(1, data.capacity), created_by=user.id)
    db.add(c); db.commit(); db.refresh(c)
    return class_dict(c, db)

@app.patch("/api/staff/classes/{class_id}")
def edit_class(class_id: int, data: ClassIn, user: User = Depends(current_user), db: Session = Depends(db_session)):
    c = db.get(ClassSession, class_id)
    if not c: raise HTTPException(404, "not_found")
    if not can(user, "classes.edit"):
        if not (can(user, "classes.edit_own") and user.coach and c.coach_id == user.coach.id): raise HTTPException(403, "permission_denied")
    c.studio_id=data.studio_id; c.title=data.title; c.class_type=data.class_type; c.starts_at=data.starts_at; c.duration=data.duration; c.capacity=data.capacity
    if can(user, "classes.edit"): c.coach_id=data.coach_id
    db.commit(); return class_dict(c, db)

@app.delete("/api/staff/classes/{class_id}")
def delete_class(class_id: int, user: User = Depends(require("classes.delete")), db: Session = Depends(db_session)):
    c=db.get(ClassSession,class_id)
    if not c: raise HTTPException(404,"not_found")
    c.status="cancelled"; db.commit(); return {"ok":True}

@app.get("/api/staff/bookings")
def staff_bookings(user: User = Depends(require("bookings.view")), db: Session = Depends(db_session)):
    rows = db.scalars(select(Booking).order_by(Booking.created_at.desc()).limit(500)).all()
    if user.coach and not can(user, "bookings.manage"):
        rows = [b for b in rows if b.klass.coach_id == user.coach.id]
    imported_total = db.scalar(select(func.coalesce(func.sum(ClassSession.source_bookings_total), 0))) or 0
    return {
        "bookings": [{"id":b.id,"reference":b.reference,"class_id":b.class_id,"class_name":b.klass.title,"starts_at":b.klass.starts_at.isoformat(),"studio":b.klass.studio.name,"customer_name":b.customer_name,"email":b.email,"phone":b.phone,"spot_number":b.spot_number,"status":b.status,"payment_status":b.payment_status,"payment_method":b.payment_method,"amount_cents":b.amount_cents} for b in rows],
        "imported_booking_count": int(imported_total),
        "imported_booking_mode": "aggregated_without_personal_data",
    }

@app.patch("/api/staff/bookings/{booking_id}")
def staff_booking_update(booking_id: int, data: BookingUpdate, user: User = Depends(require("bookings.manage")), db: Session = Depends(db_session)):
    b=db.get(Booking,booking_id)
    if not b: raise HTTPException(404,"not_found")
    if data.status is not None: b.status=data.status
    if data.payment_status is not None: b.payment_status=data.payment_status
    if data.amount_cents is not None: b.amount_cents=max(0,data.amount_cents)
    db.commit(); return {"ok":True}

def parse_upload_rows(content: bytes, filename: str):
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        return list(csv.DictReader(io.StringIO(text)))
    if ext in (".xlsx", ".xlsm"):
        wb=load_workbook(io.BytesIO(content), read_only=True, data_only=True); ws=wb.active
        rows=list(ws.iter_rows(values_only=True));
        if not rows: return []
        headers=[str(x or "").strip() for x in rows[0]]
        return [{headers[i]: row[i] for i in range(min(len(headers),len(row)))} for row in rows[1:] if any(v is not None for v in row)]
    return []

def pick(row, *names):
    normalized={str(k).strip().lower():v for k,v in row.items()}
    for n in names:
        if n.lower() in normalized and normalized[n.lower()] not in (None,""): return normalized[n.lower()]
    return None

def studio_from_value(value, db):
    s=str(value or "").lower()
    for sid,name,_,_ in STUDIOS:
        if sid in s or name.lower().split(" · ")[0] in s: return sid
    return None

@app.get("/api/staff/uploads")
def list_uploads(user: User = Depends(require("schedules.upload")), db: Session = Depends(db_session)):
    rows=db.scalars(select(ScheduleUpload).order_by(ScheduleUpload.created_at.desc()).limit(100)).all()
    return {"uploads":[{"id":x.id,"filename":x.filename,"status":x.status,"imported_rows":x.imported_rows,"created_at":x.created_at.isoformat()} for x in rows]}

@app.post("/api/staff/uploads/schedule")
async def upload_schedule(file: UploadFile = File(...), user: User = Depends(require("schedules.upload")), db: Session = Depends(db_session)):
    content=await file.read()
    if len(content)>10*1024*1024: raise HTTPException(413,"file_too_large")
    ext=Path(file.filename or "upload").suffix.lower()
    if ext not in (".csv",".xlsx",".xlsm",".pdf"): raise HTTPException(400,"unsupported_file")
    safe_name=f"{datetime.now().strftime('%Y%m%d-%H%M%S')}-{secrets.token_hex(4)}{ext}"
    path=UPLOAD_DIR/safe_name; path.write_bytes(content)
    record=ScheduleUpload(filename=file.filename or safe_name,saved_path=str(path),uploaded_by=user.id,status="received")
    db.add(record); db.flush(); imported=0
    if ext != ".pdf":
        try:
            for row in parse_upload_rows(content,file.filename or safe_name):
                sid=studio_from_value(pick(row,"studio","standort","location"),db)
                title=str(pick(row,"kurs","class","title","name") or "Class").strip()
                ctype=str(pick(row,"type","class type","typ") or "Reformer").strip()
                datev=pick(row,"datum","date","tag"); timev=pick(row,"uhrzeit","time","start"); combined=pick(row,"starts_at","start at")
                if combined: starts=date_parser.parse(str(combined))
                elif datev and timev: starts=date_parser.parse(f"{datev} {timev}")
                else: continue
                coach_id=user.coach.id if user.coach and not can(user,"classes.edit") else None
                coach_name=str(pick(row,"coach","trainer","instructor") or "").strip()
                if coach_id is None and coach_name:
                    coach=db.scalar(select(Coach).where(func.lower(Coach.display_name)==coach_name.lower())); coach_id=coach.id if coach else None
                cap=int(pick(row,"capacity","plätze","plaetze","anzahl der plätze") or (db.get(Studio,sid).capacity if sid and db.get(Studio,sid) else 10))
                duration=int(pick(row,"duration","dauer") or 50)
                if not sid: continue
                db.add(ClassSession(studio_id=sid,title=title,class_type=ctype,coach_id=coach_id,starts_at=starts,duration=duration,capacity=cap,created_by=user.id)); imported+=1
            record.status="imported" if imported else "received_no_rows"; record.imported_rows=imported
        except Exception as exc:
            record.status="received_parse_error"
    else:
        record.status="received_pdf"
    db.commit(); return {"ok":True,"status":record.status,"imported_rows":imported}

@app.get("/api/staff/marketing")
def marketing(user: User = Depends(require("pro.view"))):
    return {"locked": True, "premium": True, "title": "Email Marketing", "message": "Premium module — available as a future upgrade."}

@app.get("/api/public/overview")
def public_overview(db: Session = Depends(db_session)):
    latest = db.scalar(select(func.max(ClassSession.starts_at)))
    earliest = db.scalar(select(func.min(ClassSession.starts_at)))
    return {
        "studios": db.scalar(select(func.count(Studio.id))) or 0,
        "coaches": db.scalar(select(func.count(Coach.id)).where(Coach.active == True)) or 0,
        "class_formats": db.scalar(select(func.count(func.distinct(ClassSession.title)))) or 0,
        "sessions": db.scalar(select(func.count(ClassSession.id))) or 0,
        "bookings": int(db.scalar(select(func.coalesce(func.sum(ClassSession.source_bookings_total), 0))) or 0),
        "available_from": earliest.isoformat() if earliest else None,
        "available_to": latest.isoformat() if latest else None,
        "source": "Mindbody export",
    }


@app.get("/api/public/coaches")
def public_coaches(db: Session = Depends(db_session)):
    session_counts = dict(db.execute(
        select(ClassSession.coach_id, func.count(ClassSession.id))
        .where(ClassSession.coach_id.is_not(None))
        .group_by(ClassSession.coach_id)
    ).all())
    booking_counts = dict(db.execute(
        select(ClassSession.coach_id, func.coalesce(func.sum(ClassSession.source_bookings_total), 0))
        .where(ClassSession.coach_id.is_not(None))
        .group_by(ClassSession.coach_id)
    ).all())
    coach_studios: dict[int, set[str]] = {}
    for coach_id, studio_name in db.execute(
        select(ClassSession.coach_id, Studio.name)
        .join(Studio, Studio.id == ClassSession.studio_id)
        .where(ClassSession.coach_id.is_not(None))
        .distinct()
    ).all():
        coach_studios.setdefault(coach_id, set()).add(studio_name)
    coaches = db.scalars(select(Coach).where(Coach.active == True).order_by(Coach.display_name)).all()
    return {"coaches": [{
        "id": coach.id,
        "display_name": coach.display_name,
        "photo_url": coach.photo_url,
        "bio": coach.bio,
        "sessions": int(session_counts.get(coach.id, 0)),
        "bookings": int(booking_counts.get(coach.id, 0)),
        "studios": sorted(coach_studios.get(coach.id, set())),
    } for coach in coaches]}


@app.get("/api/public/classes")
def public_class_catalog(db: Session = Depends(db_session)):
    rows = db.execute(
        select(
            ClassSession.title,
            ClassSession.class_type,
            Studio.id,
            Studio.name,
            func.count(ClassSession.id),
            func.coalesce(func.sum(ClassSession.source_bookings_total), 0),
            func.count(func.distinct(ClassSession.coach_id)),
        )
        .join(Studio, Studio.id == ClassSession.studio_id)
        .group_by(ClassSession.title, ClassSession.class_type, Studio.id, Studio.name)
        .order_by(ClassSession.title)
    ).all()
    return {"classes": [{
        "name": title, "type": class_type, "studio": studio_id, "studio_name": studio_name,
        "sessions": int(sessions), "bookings": int(bookings), "coaches": int(coaches),
    } for title, class_type, studio_id, studio_name, sessions, bookings, coaches in rows]}


def berlin_day(value: str, *, end: bool = False) -> datetime:
    parsed = date.fromisoformat(value)
    if end:
        parsed += timedelta(days=1)
    return datetime.combine(parsed, time.min, tzinfo=ZoneInfo("Europe/Berlin"))


@app.get("/api/schedule")
def public_schedule(
    from_: Optional[str] = Query(default=None, alias="from"),
    to: Optional[str] = Query(default=None),
    db: Session = Depends(db_session),
):
    try:
        start = berlin_day(from_) if from_ else datetime.now(ZoneInfo("Europe/Berlin")).replace(hour=0, minute=0, second=0, microsecond=0)
        end = berlin_day(to, end=True) if to else start + timedelta(days=14)
    except ValueError:
        raise HTTPException(422, "invalid_date_range")
    if end <= start or end - start > timedelta(days=45):
        raise HTTPException(422, "invalid_date_range")
    now = datetime.now(timezone.utc)
    visible_from = max(as_utc(start), now)
    q=select(ClassSession).where(
        ClassSession.starts_at >= visible_from,
        ClassSession.starts_at < end,
        ClassSession.status == "active",
    ).order_by(ClassSession.starts_at).limit(1200)
    rows=db.scalars(q).all()
    return {"classes":[class_dict(c,db) for c in rows], "from": from_, "to": to}

FALLBACK_COACHES = {"Anna", "Andrea", "Christina", "Gabriella", "Ida", "Jessi", "Kimberley", "Melina", "Nathalie", "Sani", "Zora", "Laetitia"}
FALLBACK_CLASS_TYPES = {"Reformer", "Powerformer", "Mat", "Barre"}

def resolve_public_class(data: PublicBookingIn, user: Optional[User], db: Session):
    try:
        class_id = int(data.classId)
    except (TypeError, ValueError):
        class_id = None
    if class_id is not None:
        return db.get(ClassSession, class_id)

    external_id = str(data.classId).strip()
    mapped = db.get(PublicClassMap, external_id)
    if mapped:
        return db.get(ClassSession, mapped.class_id)
    if not user or portal_for(user) != "/account":
        raise HTTPException(401, "customer_login_required")
    match = re.fullmatch(r"(\d{4}-\d{2}-\d{2})-(bhf1|ladies|sachsen|bornheim|mid|oval)-(\d{4})", external_id)
    if not match or data.studioId != match.group(2) or data.classType not in FALLBACK_CLASS_TYPES or not data.startsAt:
        raise HTTPException(400, "invalid_public_class")
    starts_at = as_utc(data.startsAt)
    now = datetime.now(timezone.utc)
    if starts_at < now - timedelta(hours=1) or starts_at > now + timedelta(days=21):
        raise HTTPException(400, "invalid_public_class_time")
    if not (2 <= len(data.title.strip()) <= 180):
        raise HTTPException(400, "invalid_public_class_title")
    coach_id = None
    coach_name = data.coachName.strip()
    if coach_name in FALLBACK_COACHES:
        coach = db.scalar(select(Coach).where(func.lower(Coach.display_name) == coach_name.lower()))
        if not coach:
            coach = Coach(display_name=coach_name)
            db.add(coach); db.flush()
        coach_id = coach.id
    klass = ClassSession(
        studio_id=data.studioId,
        title=data.title.strip(),
        class_type=data.classType,
        coach_id=coach_id,
        starts_at=starts_at,
        duration=min(120, max(15, data.duration)),
        capacity=min(30, max(1, data.capacity)),
        created_by=user.id,
    )
    db.add(klass); db.flush()
    db.add(PublicClassMap(external_id=external_id, class_id=klass.id)); db.flush()
    return klass

@app.post("/api/bookings")
def public_booking(data: PublicBookingIn, user: Optional[User] = Depends(optional_user), db: Session = Depends(db_session)):
    c=resolve_public_class(data,user,db)
    if not c or c.status!="active": raise HTTPException(409,"class_unavailable")
    if as_utc(c.starts_at) <= datetime.now(timezone.utc): raise HTTPException(409,"class_started")
    live_reserved=db.scalar(select(func.count(Booking.id)).where(Booking.class_id==c.id,Booking.status=="reserved")) or 0
    reserved=int(c.imported_bookings or 0)+int(live_reserved)
    if reserved>=c.capacity: raise HTTPException(409,"class_full")
    duplicate=db.scalar(select(Booking).where(Booking.class_id==c.id,Booking.email==data.email.lower(),Booking.status=="reserved"))
    if duplicate: raise HTTPException(409,"duplicate_booking")
    if data.spot:
        if data.spot <= int(c.imported_bookings or 0): raise HTTPException(409,"spot_taken")
        spot_taken=db.scalar(select(Booking).where(Booking.class_id==c.id,Booking.spot_number==data.spot,Booking.status=="reserved"))
        if spot_taken: raise HTTPException(409,"spot_taken")
    ref="CP-"+secrets.token_hex(4).upper()
    b=Booking(reference=ref,class_id=c.id,customer_name=(data.firstName+" "+data.lastName).strip(),email=data.email.lower(),phone=data.phone,spot_number=data.spot,payment_method=data.paymentMethod,payment_status="pending",amount_cents=2800 if data.paymentMethod else 0)
    db.add(b); db.flush()
    if user and portal_for(user) == "/account" and user.email.lower() == data.email.lower():
        db.add(CustomerBookingLink(booking_id=b.id, user_id=user.id))
    db.commit();return {"booking":{"reference":ref},"payment_status":b.payment_status}

@app.get("/api/bookings")
def public_bookings(email: str, reference: str, db: Session = Depends(db_session)):
    normalized_ref = reference.strip().upper()
    booking = db.scalar(select(Booking).where(Booking.email == email.strip().lower(), Booking.reference == normalized_ref))
    if not booking:
        raise HTTPException(404, "booking_not_found")
    return {"credits":0,"bookings":[{"reference":booking.reference,"status":booking.status,"starts_at":booking.klass.starts_at.isoformat(),"name":booking.klass.title,"studio_name":booking.klass.studio.name,"spot_number":booking.spot_number}]}

@app.delete("/api/bookings")
def public_cancel(payload: dict, db: Session = Depends(db_session)):
    ref=str(payload.get("reference", "")); email=str(payload.get("email", "")).lower()
    b=db.scalar(select(Booking).where(Booking.reference==ref,Booking.email==email))
    if not b: raise HTTPException(404,"not_found")
    b.status="cancelled";db.commit();return {"ok":True}

@app.post("/api/waitlist")
def join_waitlist(data: WaitlistIn, user: Optional[User] = Depends(optional_user), db: Session = Depends(db_session)):
    c=db.get(ClassSession,data.classId)
    if not c: raise HTTPException(404,"not_found")
    live_reserved=db.scalar(select(func.count(Booking.id)).where(Booking.class_id==c.id,Booking.status=="reserved")) or 0
    reserved=int(c.imported_bookings or 0)+int(live_reserved)
    if reserved<c.capacity: raise HTTPException(409,"spots_available")
    existing=db.scalar(select(Waitlist).where(Waitlist.class_id==c.id,Waitlist.email==data.email.lower()))
    if existing: raise HTTPException(409,"already_waitlisted")
    w=Waitlist(class_id=c.id,email=data.email.lower());db.add(w);db.flush()
    if user and portal_for(user) == "/account" and user.email.lower() == data.email.lower():
        db.add(CustomerWaitlistLink(waitlist_id=w.id, user_id=user.id))
    db.commit()
    pos=db.scalar(select(func.count(Waitlist.id)).where(Waitlist.class_id==c.id,Waitlist.created_at<=w.created_at)) or 1
    return {"position":pos}
